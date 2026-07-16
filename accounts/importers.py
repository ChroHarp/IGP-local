from dataclasses import dataclass, field
from datetime import date, datetime
from typing import BinaryIO
from zipfile import BadZipFile

from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .models import AwardRecord, FamilyMember, Guardian, InitialIGPProfile, Student


class StudentImportError(ValueError):
    pass


@dataclass
class StudentImportResult:
    row_count: int = 0
    create_count: int = 0
    skip_count: int = 0
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self):
        return not self.errors

    @property
    def error_count(self):
        return len(self.errors)


HEADER_ALIASES = {
    "student_number": {"學號", "學生學號"},
    "full_name": {"學生姓名"},
    "gender": {"性別"},
    "grade": {"年級"},
    "class_name": {"班別", "班級"},
    "seat_number": {"座號"},
    "date_of_birth": {"出生年月日", "出生日期"},
    "email": {"學生email", "學生e-mail"},
    "home_phone": {"住家電話"},
    "address": {"住址"},
    "guardian_name": {"法定代理人"},
    "guardian_work": {"法定代理人連絡電話(公司)", "法定代理人聯絡電話(公司)"},
    "guardian_mobile": {"法定代理人連絡電話(手機)", "法定代理人聯絡電話(手機)"},
    "multiple_needs": {"是否有雙重特教需求"},
}

PROFILE_HEADERS = {
    "source_submitted_at": ("時間戳記",),
    "source_email": ("電子郵件地址",),
    "additional_family_notes": ("若尚有其他成員，請依上述格式條列於此",),
    "family_culture": ("家庭文化特質",),
    "primary_caregiver": ("主要照顧者",),
    "learning_supporter": ("主要協助學習者",),
    "household_economy": ("家庭經濟狀況",),
    "caregiving_style": ("照顧者管教態度",),
    "family_interaction": ("與家人互動情形",),
    "math_aptitude_score": ("第一階段-數學性向測驗分數",),
    "science_aptitude_score": ("第一階段-自然性向測驗分數",),
    "math_practical_t_score": ("第二階段-數學實作評量t分數",),
    "science_practical_t_score": ("第二階段-自然實作評量t分數",),
    "science_interests": ("科學興趣",),
    "arts_interests": ("人文與藝術興趣",),
    "other_interests": ("其他興趣",),
    "other_awards_notes": ("其他得獎紀錄請依上述格式條列於此",),
    "completed_by": ("填寫者",),
    "cognitive_strengths": ("認知優勢特質",),
    "emotional_strengths": ("情意優勢特質",),
    "academic_strengths": ("學科優勢能力",),
    "cognitive_needs": ("認知弱勢特質",),
    "emotional_needs": ("情意弱勢特質",),
    "academic_needs": ("學科弱勢能力",),
    "notes": ("其他補充說明事項",),
}


def normalize_header(value):
    return "".join(str(value or "").lower().split())


def normalize_text(value):
    return str(value).strip() if value is not None else ""


def as_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = normalize_text(value).replace("/", "-")
    try:
        return date.fromisoformat(text)
    except ValueError as exc:
        raise StudentImportError(f"無法辨識出生日期「{value}」。") from exc


def as_int(value, label):
    if value in (None, ""):
        return None
    try:
        numeric = int(float(value))
    except (TypeError, ValueError) as exc:
        raise StudentImportError(f"{label} 必須是整數。") from exc
    if numeric < 0:
        raise StudentImportError(f"{label} 不可為負數。")
    return numeric


def as_gender(value):
    text = normalize_text(value)
    mapping = {"女": Student.Gender.FEMALE, "男": Student.Gender.MALE}
    return mapping.get(text, Student.Gender.OTHER if text else Student.Gender.UNSPECIFIED)


def as_bool(value):
    return normalize_text(value).lower() in {"是", "yes", "true", "1", "有"}


def header_indexes(header_row):
    normalized = {normalize_header(value): index for index, value in enumerate(header_row)}
    indexes = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            index = normalized.get(normalize_header(alias))
            if index is not None:
                indexes[field] = index
                break
    if "full_name" not in indexes:
        raise StudentImportError("找不到必要欄位「學生姓名」。")
    return indexes


def raw_response(headers, row):
    result = {}
    for header, value in zip(headers, row):
        key = normalize_text(header)
        if not key:
            continue
        candidate = key
        suffix = 2
        while candidate in result:
            candidate = f"{key}__{suffix}"
            suffix += 1
        result[candidate] = normalize_text(value)
    return result


def response_value(response, *headers):
    wanted = {normalize_header(header) for header in headers}
    for key, value in response.items():
        if normalize_header(key) in wanted:
            return normalize_text(value)
    return ""


def response_key(response, key):
    return normalize_text(response.get(key, ""))


def profile_values(response):
    return {
        field: response_value(response, *headers)
        for field, headers in PROFILE_HEADERS.items()
    }


def create_family_members(student, response):
    rows = [
        ("父親", response_value(response, "父親姓名"), response_value(response, "父親服務機關"), response_value(response, "父親畢業科系"), response_value(response, "父親專長"), response_value(response, "父親連絡電話")),
        ("母親", response_value(response, "母親姓名"), response_value(response, "母親-服務機關"), response_value(response, "母親-畢業科系"), response_value(response, "母親-專長"), response_value(response, "母親-連絡電話")),
        (response_key(response, "稱謂"), response_value(response, "家人1-姓名"), response_value(response, "家人1-服務機關/就讀學校"), response_value(response, "家人1-畢業科系 (在學免填)"), response_value(response, "家人1-專長"), response_value(response, "家人1-連絡電話")),
        (response_key(response, "稱謂__2"), response_value(response, "家人2-姓名"), response_value(response, "家人2-服務機關/就讀學校"), response_value(response, "家人2-畢業科系(在學免填)"), response_value(response, "家人2-專長"), response_value(response, "家人2-連絡電話")),
        (response_key(response, "稱謂__3"), response_value(response, "家人3-姓名"), response_value(response, "家人3-服務機關/就讀學校"), response_value(response, "家人3-畢業科系 (在學免填)"), response_value(response, "家人3-專長"), response_value(response, "家人3-連絡電話")),
    ]
    for sort_order, (relationship, full_name, organization, education, specialty, phone) in enumerate(rows):
        if not full_name:
            continue
        FamilyMember.objects.get_or_create(
            student=student,
            relationship=relationship or "其他",
            full_name=full_name,
            defaults={
                "organization_or_school": organization,
                "education_major": education,
                "specialty": specialty,
                "phone": phone,
                "sort_order": sort_order,
            },
        )


def create_award_record(student, response):
    values = {
        "award_date": response_value(response, "獲獎日期"),
        "activity_name": response_value(response, "競賽/活動名稱"),
        "organizer": response_value(response, "主辦單位"),
        "award": response_value(response, "獎項"),
        "award_type": response_value(response, "得獎類型"),
    }
    if not any(values.values()):
        return
    AwardRecord.objects.get_or_create(student=student, **values)


def populate_imported_details(student, response, *, overwrite=False):
    profile, _ = InitialIGPProfile.objects.get_or_create(student=student, defaults={"raw_response": response})
    values = profile_values(response)
    for field, value in values.items():
        if overwrite or not getattr(profile, field):
            setattr(profile, field, value)
    if response:
        profile.raw_response = response
    profile.save()

    if overwrite or not student.has_multiple_special_education_needs:
        student.has_multiple_special_education_needs = as_bool(response_value(response, "是否有雙重特教需求"))
        student.save(update_fields=["has_multiple_special_education_needs"])
    create_family_members(student, response)
    create_award_record(student, response)


def cell(row, indexes, field):
    index = indexes.get(field)
    return row[index] if index is not None and index < len(row) else None


def parse_students(workbook_file: BinaryIO):
    try:
        workbook = load_workbook(workbook_file, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise StudentImportError("Excel 檔案格式不正確或已損毀。") from exc
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = tuple(next(rows))
        indexes = header_indexes(headers)
    except StopIteration as exc:
        raise StudentImportError("Excel 沒有資料列。") from exc

    parsed = []
    errors = []
    for row_number, row in enumerate(rows, start=2):
        full_name = normalize_text(cell(row, indexes, "full_name"))
        if not full_name:
            continue
        try:
            parsed.append({
                "row_number": row_number,
                "student_number": normalize_text(cell(row, indexes, "student_number")) or None,
                "full_name": full_name,
                "gender": as_gender(cell(row, indexes, "gender")),
                "has_multiple_special_education_needs": as_bool(cell(row, indexes, "multiple_needs")),
                "date_of_birth": as_date(cell(row, indexes, "date_of_birth")),
                "grade": as_int(cell(row, indexes, "grade"), "年級"),
                "class_name": normalize_text(cell(row, indexes, "class_name")),
                "seat_number": as_int(cell(row, indexes, "seat_number"), "座號"),
                "email": normalize_text(cell(row, indexes, "email")),
                "home_phone": normalize_text(cell(row, indexes, "home_phone")),
                "address": normalize_text(cell(row, indexes, "address")),
                "guardian_name": normalize_text(cell(row, indexes, "guardian_name")),
                "guardian_work": normalize_text(cell(row, indexes, "guardian_work")),
                "guardian_mobile": normalize_text(cell(row, indexes, "guardian_mobile")),
                "raw_response": raw_response(headers, row),
            })
        except StudentImportError as exc:
            errors.append(f"第 {row_number} 列：{exc}")
    return parsed, errors

def existing_student(data):
    if data["student_number"]:
        return Student.objects.filter(student_number=data["student_number"]).first()
    candidates = Student.objects.filter(full_name=data["full_name"], date_of_birth=data["date_of_birth"])
    if candidates.count() > 1:
        raise StudentImportError("同名且出生日期相同，請先補學號後再匯入。")
    return candidates.first()


def import_basic_students(workbook_file: BinaryIO, *, apply=False):
    result = StudentImportResult()
    try:
        parsed, parse_errors = parse_students(workbook_file)
    except StudentImportError as exc:
        result.errors.append(str(exc))
        return result

    result.row_count = len(parsed) + len(parse_errors)
    result.errors.extend(parse_errors)

    pending = []
    seen_identities = set()
    for data in parsed:
        try:
            identity = ("student_number", data["student_number"]) if data["student_number"] else ("name_birth", data["full_name"], data["date_of_birth"])
            if identity in seen_identities:
                raise StudentImportError("Excel 內有重複學生；請補學號或合併重複列。")
            seen_identities.add(identity)
            if existing_student(data):
                result.skip_count += 1
                continue
            student = Student(
                student_number=data["student_number"],
                full_name=data["full_name"],
                gender=data["gender"],
                has_multiple_special_education_needs=data["has_multiple_special_education_needs"],
                date_of_birth=data["date_of_birth"],
                grade=data["grade"],
                class_name=data["class_name"],
                seat_number=data["seat_number"],
                email=data["email"],
                home_phone=data["home_phone"],
                address=data["address"],
            )
            student.full_clean()
            pending.append((student, data))
            result.create_count += 1
        except StudentImportError as exc:
            result.errors.append(f"第 {data['row_number']} 列：{exc}")
        except Exception as exc:
            result.errors.append(f"第 {data['row_number']} 列：{exc}")

    if apply and result.is_valid:
        with transaction.atomic():
            for student, data in pending:
                student.save()
                populate_imported_details(student, data["raw_response"], overwrite=True)
                if data["guardian_name"]:
                    Guardian.objects.create(
                        student=student,
                        full_name=data["guardian_name"],
                        phone_work=data["guardian_work"],
                        phone_mobile=data["guardian_mobile"],
                        is_primary=True,
                    )
    return result
